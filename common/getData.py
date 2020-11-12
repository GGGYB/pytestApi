import openpyxl
from openpyxl import workbook
import sqlite3
import pprint
import json
def get_excel_data(sheetName,startRow,endRow,urlColumn,bodyColumn,expColumn):
    fileDir = r"../data/pytestApi.xlsx"
    wb = openpyxl.load_workbook(fileDir)
    ws = wb.active
    dataList = []
    for i in range(startRow,endRow):
        urlCell = ws.cell(row=startRow,column=urlColumn).value
        bodyCell = ws.cell(row=startRow,column =bodyColumn ).value
        expCell = ws.cell(row=startRow,column =expColumn).value
        dataList.append([urlCell.strip(),bodyCell.strip(),expCell.strip()])
    # print(dataList)
    return dataList

def get_db_data(db,table):
    conn = sqlite3.connect(db)
    cursor = conn.cursor()
    sql = f'select * from {table}'
    cursor.execute(sql)
    values = cursor.fetchall()
    pprint.pprint(values)
    conn.commit()
    cursor.close()
    conn.close()
    return values

# a = get_excel_data("user",2,6,4,7,8)
# b = type(a[0][1])
# print(b)
# print(a[0][1])
# print(json.loads(b))

get_db_data("../data/user","user")